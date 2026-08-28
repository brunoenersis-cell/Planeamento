"""Exportações Excel e CSV."""
from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_csv(frame: pd.DataFrame) -> bytes:
    return frame.to_csv(index=False).encode("utf-8-sig")


def create_excel(frame: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="Cronograma")
        frame.to_excel(writer, index=False, sheet_name="Dados")
        status = frame["Estado"].fillna("").astype(str).str.upper() if "Estado" in frame else pd.Series(dtype=str)
        summary = pd.DataFrame({"Indicador": ["Total de atividades", "Concluídas", "Pendentes", "Futuras"], "Valor": [len(frame), int(status.isin(["COMPLETE", "LATE_COMPLETE", "COMPLETED"]).sum()), int(status.isin(["TODO", "MISSED", "PENDING"]).sum()), int((pd.to_datetime(frame.get("Data"), errors="coerce", utc=True) > pd.Timestamp.now(tz="UTC")).sum())]})
        summary.to_excel(writer, index=False, sheet_name="Resumo")
        for sheet_name in ("Cronograma", "Dados", "Resumo"):
            sheet = writer.sheets[sheet_name]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for column in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 40)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = width
            if sheet_name in ("Cronograma", "Dados") and len(frame) > 0:
                table = Table(displayName=f"Tabela{sheet_name}", ref=sheet.dimensions)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                sheet.add_table(table)
    return output.getvalue()

